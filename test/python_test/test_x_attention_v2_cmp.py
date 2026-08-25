#!/usr/bin/env python3
"""XAttention v1 vs v2 three-way precision comparison.

For each of 144 test cases:
1. Load CPU golden from golden cache (shared with test_x_attention.py)
2. Run v1 (XAttention) and v2 (XAttentionV2) on NPU with identical inputs
3. Three-way compare: v2_output vs v1_output vs CPU_golden
   using inlined precision_compare_triple_data (from PyPTO_Test/libs/precision_compare.py)
4. Collect per-case metrics (MARE/MERE/RMSE for both, ratios, judgment)

After all cases complete, write results to xlsx:
  test_x_attention_v2_cmp/x_attention_v2_cmp_rst_<timestamp>.xlsx

Usage:
    pytest test_x_attention_v2_cmp.py -v

Prerequisite:
    Golden cache must exist (run test_x_attention.py first to generate).
"""

import os
import copy
import pytest
import torch
import numpy as np
import ml_dtypes
import statistics
from enum import Enum
from datetime import datetime
from openpyxl import Workbook

torch_npu = pytest.importorskip("torch_npu")
custom_ops = pytest.importorskip("custom_ops")

WORKSPACE = os.path.dirname(os.path.abspath(__file__))
GOLDEN_CACHE_DIR = os.path.join(WORKSPACE, "golden_cache")
OUTPUT_DIR = os.path.join(WORKSPACE, "test_x_attention_v2_cmp")

_DTYPE_NAME_MAP = {torch.bfloat16: "bf16", torch.float16: "fp16"}


# ============================================================
# Inlined precision compare (minimal extraction from
# PyPTO_Test/libs/precision_compare.py + ReturnValue.py)
# ============================================================

class Result(Enum):
    PASS = "PASS"
    FAILED = "FAILED"
    ERROR = "ERROR"
    IGNORE = "IGNORE"
    INVALID = "INVALID"


small_value_thres_dict = {
    "fp16": 2 ** -11, "float16": 2 ** -11,
    "bf16": 2 ** -8, "bfloat16": 2 ** -8,
    "fp32": 2 ** -14, "float32": 2 ** -14,
}

small_value_error_thres_dict = {
    "fp16": 2 ** -16, "float16": 2 ** -16,
    "bf16": 2 ** -16, "bfloat16": 2 ** -16,
    "fp32": 2 ** -30, "float32": 2 ** -30,
}


def _filter_inf_nan(input_data, golden_data, bm_data=None):
    if np.issubdtype(input_data.dtype, np.integer):
        if bm_data is not None:
            return input_data, golden_data, bm_data
        else:
            return input_data, golden_data

    if input_data.dtype in [np.float16, ml_dtypes.bfloat16]:
        if input_data.dtype in [np.float16]:
            precision_index = np.where(np.abs(golden_data) < 65504)[0]
        else:
            precision_index = np.where(np.abs(golden_data) < 3.38953e38)[0]
    else:
        precision_index = None
    inf_index = np.where(np.logical_not(np.isinf(golden_data)))[0]
    nan_index = np.where(np.logical_not(np.isnan(golden_data)))[0]
    filter_index = np.intersect1d(inf_index, nan_index)
    if precision_index is not None:
        filter_index = np.intersect1d(filter_index, precision_index)

    input_data_filter = input_data[filter_index]
    golden_data_filter = golden_data[filter_index]
    if bm_data is not None:
        bm_data_filter = bm_data[filter_index]
        return input_data_filter, golden_data_filter, bm_data_filter
    else:
        return input_data_filter, golden_data_filter


def _get_split_index(golden_data, dtype):
    small_value_thres = small_value_thres_dict[dtype]
    large_value_idx = np.where(np.abs(golden_data) >= small_value_thres)[0]
    small_value_idx = np.where(np.abs(golden_data) < small_value_thres)[0]
    return large_value_idx, small_value_idx, small_value_thres


def _compute_matrix_small_value(input_data, golden_data, dtype, small_index):
    if small_index.size == 0:
        return 0
    thres = small_value_error_thres_dict[dtype]
    error_idx = np.where(np.abs(input_data[small_index] - golden_data[small_index]) > thres)[0]
    return error_idx.size


def _compute_matrix_large_value(input_data, golden_data, large_index):
    if large_index.size == 0:
        return 0.0, 0.0, 0.0
    input_data_large = input_data[large_index]
    golden_data_large = golden_data[large_index]
    relative_error = np.abs(input_data_large - golden_data_large) / (np.abs(golden_data_large) + 1e-7)
    mare = float(np.max(relative_error))
    mere = float(np.mean(relative_error))
    rmse = float(np.sqrt(np.mean((input_data_large - golden_data_large) ** 2)))
    return mare, mere, rmse


def _compute_re_matrix(input_value, bm_value, small_value_thres):
    if np.isinf(bm_value) or np.isnan(bm_value):
        return 1.0
    if np.isinf(input_value) or np.isnan(input_value):
        return 1000.0
    return input_value / max(bm_value, small_value_thres)


def precision_compare_triple_data(npu_data, bm_data, golden_data, dtype, thres=(2, 1.2, 1.2)):
    """Three-way precision compare: npu_data(v2) vs bm_data(v1) vs golden_data(CPU).

    Ratios are v2_error / v1_error. <1 means v2 is more precise.

    Returns dict with:
        result: Result.PASS / Result.FAILED / Result.INVALID
        mare_ratio, mere_ratio, rmse_ratio, small_ratio: float (v2/v1)
        v2_mare, v2_mere, v2_rmse, v2_small_err: float (v2 absolute metrics)
        v1_mare, v1_mere, v1_rmse, v1_small_err: float (v1 absolute metrics)
        small_total, large_total: int (element counts)
    """
    if dtype in ["int8", "int32"]:
        raise NotImplementedError("precision compare triplet only support float")
    npu_data = npu_data.flatten().astype(np.float32)
    bm_data = bm_data.flatten().astype(np.float32)
    golden_data = golden_data.flatten().astype(np.float32)

    npu_data, golden_data, bm_data = _filter_inf_nan(npu_data, golden_data, bm_data)
    if golden_data.size == 0:
        return {"result": Result.INVALID,
                "mare_ratio": 0.0, "mere_ratio": 0.0, "rmse_ratio": 0.0, "small_ratio": 0.0,
                "v2_mare": 0.0, "v2_mere": 0.0, "v2_rmse": 0.0, "v2_small_err": 0,
                "v1_mare": 0.0, "v1_mere": 0.0, "v1_rmse": 0.0, "v1_small_err": 0,
                "small_total": 0, "large_total": 0}

    large_value_idx, small_value_idx, small_value_thres = _get_split_index(golden_data, dtype)

    npu_small_err = _compute_matrix_small_value(npu_data, golden_data, dtype, small_value_idx)
    bm_small_err = _compute_matrix_small_value(bm_data, golden_data, dtype, small_value_idx)
    small_value_matrix = npu_small_err / max(bm_small_err, 1)

    mare_npu, mere_npu, rmse_npu = _compute_matrix_large_value(npu_data, golden_data, large_value_idx)
    mare_bm, mere_bm, rmse_bm = _compute_matrix_large_value(bm_data, golden_data, large_value_idx)

    mare_matrix = _compute_re_matrix(mare_npu, mare_bm, small_value_thres)
    mere_matrix = _compute_re_matrix(mere_npu, mere_bm, small_value_thres)
    rmse_matrix = _compute_re_matrix(rmse_npu, rmse_bm, small_value_thres)

    if small_value_matrix <= 2 and mare_matrix <= thres[0] and mere_matrix <= thres[1] and rmse_matrix <= thres[2]:
        result = Result.PASS
    else:
        result = Result.FAILED

    return {
        "result": result,
        "mare_ratio": mare_matrix, "mere_ratio": mere_matrix,
        "rmse_ratio": rmse_matrix, "small_ratio": small_value_matrix,
        "v2_mare": mare_npu, "v2_mere": mere_npu, "v2_rmse": rmse_npu,
        "v2_small_err": npu_small_err,
        "v1_mare": mare_bm, "v1_mere": mere_bm, "v1_rmse": rmse_bm,
        "v1_small_err": bm_small_err,
        "small_total": small_value_idx.size, "large_total": large_value_idx.size,
    }


# ============================================================
# Golden cache helpers
# ============================================================

def _golden_cache_key(dtype, num_head, kv_heads, request_num, beam_size, kv_seqlen, unshared_seqlen):
    dtype_name = _DTYPE_NAME_MAP.get(dtype, str(dtype))
    return f"{dtype_name}_{num_head}_{kv_heads}_{request_num}_{beam_size}_{kv_seqlen}_{unshared_seqlen}"


def _load_golden(cache_key):
    cache_path = os.path.join(GOLDEN_CACHE_DIR, f"{cache_key}.pt")
    if not os.path.exists(cache_path):
        return None
    return torch.load(cache_path, weights_only=False)


# ============================================================
# Device op call
# ============================================================

def _call_op(op_func, cached):
    q = cached["query"].npu()
    k = cached["key_cache"].npu()
    v = cached["value_cache"].npu()
    uk = cached["unshared_key"].npu()
    uv = cached["unshared_value"].npu()

    bt = None
    if cached.get("block_tables") is not None:
        bt = torch.tensor(copy.deepcopy(cached["block_tables"]), dtype=torch.int32).npu()

    ubt = None
    if cached.get("unshared_block_tables") is not None:
        ubt = torch.tensor(copy.deepcopy(cached["unshared_block_tables"]), dtype=torch.int32).npu()

    actual_shared_kvlen = torch.tensor(cached["actual_shared_kvlen"], dtype=torch.int32).npu()
    decode_step = torch.tensor([cached["decode_step"]], dtype=torch.int32).npu()

    return op_func(q, k, v, uk, uv, actual_shared_kvlen, decode_step, bt, ubt)


# ============================================================
# Results collection and xlsx writing
# ============================================================

_RESULTS = []

_SORT_COLUMNS = ["dtype", "batch", "q_head_num", "kv_head_num",
                 "beam_size", "unshared_kv_len", "prompt_len"]

_DETAIL_HEADERS = [
    "dtype", "batch", "q_head_num", "kv_head_num", "beam_size", "prompt_len", "unshared_kv_len",
    "v1_MARE", "v1_MERE", "v1_RMSE", "v1_small_err", "v1_small_total",
    "v2_MARE", "v2_MERE", "v2_RMSE", "v2_small_err", "v2_small_total",
    "mare_ratio", "mere_ratio", "rmse_ratio", "small_ratio",
    "result",
    "v2_mare_better", "v2_mere_better", "v2_rmse_better", "v2_small_better",
]


def _auto_fit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_len = len(str(cell.value)) if cell.value is not None else 0
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[col_letter].width = max_len + 2


def _write_results_xlsx():
    if not _RESULTS:
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    xlsx_path = os.path.join(OUTPUT_DIR, f"x_attention_v2_cmp_rst_{timestamp}.xlsx")

    def _sort_key(r):
        return tuple(
            (int(r.get(c, 0)),) if str(r.get(c, "")).lstrip("-").isdigit()
            else (str(r.get(c, "")),)
            for c in _SORT_COLUMNS
        )
    sorted_results = sorted(_RESULTS, key=_sort_key)

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("detail")
    ws.append(_DETAIL_HEADERS)
    for r in sorted_results:
        ws.append([r.get(h, "") for h in _DETAIL_HEADERS])
    _auto_fit_columns(ws)

    ws2 = wb.create_sheet("summary")
    total = len(sorted_results)
    pass_count = sum(1 for r in sorted_results if r.get("result") == Result.PASS)
    fail_count = total - pass_count

    v2_mare_better = sum(1 for r in sorted_results if r.get("v2_mare_better"))
    v2_mere_better = sum(1 for r in sorted_results if r.get("v2_mere_better"))
    v2_rmse_better = sum(1 for r in sorted_results if r.get("v2_rmse_better"))
    v2_small_better = sum(1 for r in sorted_results if r.get("v2_small_better"))

    mare_ratios = [r["mare_ratio"] for r in sorted_results]
    mere_ratios = [r["mere_ratio"] for r in sorted_results]
    rmse_ratios = [r["rmse_ratio"] for r in sorted_results]
    small_ratios = [r["small_ratio"] for r in sorted_results]

    summary_rows = [
        ["Metric", "Value"],
        ["Total cases", total],
        ["PASS (v2 within threshold)", pass_count],
        ["FAILED (v2 exceeds threshold)", fail_count],
        ["", ""],
        ["v2 MARE better count (<1.0)", v2_mare_better],
        ["v2 MERE better count (<1.0)", v2_mere_better],
        ["v2 RMSE better count (<1.0)", v2_rmse_better],
        ["v2 small_value better count (<1.0)", v2_small_better],
        ["", ""],
        ["mare_ratio avg", f"{statistics.mean(mare_ratios):.6f}"],
        ["mare_ratio median", f"{statistics.median(mare_ratios):.6f}"],
        ["mere_ratio avg", f"{statistics.mean(mere_ratios):.6f}"],
        ["mere_ratio median", f"{statistics.median(mere_ratios):.6f}"],
        ["rmse_ratio avg", f"{statistics.mean(rmse_ratios):.6f}"],
        ["rmse_ratio median", f"{statistics.median(rmse_ratios):.6f}"],
        ["small_ratio avg", f"{statistics.mean(small_ratios):.6f}"],
        ["small_ratio median", f"{statistics.median(small_ratios):.6f}"],
    ]
    for row in summary_rows:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 20

    wb.save(xlsx_path)
    print(f"\n{'='*60}")
    print(f"Precision comparison results saved to:\n  {xlsx_path}")
    print(f"Total: {total}, PASS: {pass_count}, FAILED: {fail_count}")
    print(f"v2 better - MARE: {v2_mare_better}, MERE: {v2_mere_better}, "
          f"RMSE: {v2_rmse_better}, small: {v2_small_better}")
    print(f"{'='*60}")


@pytest.fixture(scope="session", autouse=True)
def _session_teardown():
    yield
    _write_results_xlsx()


# ============================================================
# Test cases (same parametrize as test_x_attention.py, 144 cases)
# ============================================================

@pytest.mark.parametrize("dtype", [torch.bfloat16])
@pytest.mark.parametrize("num_head, kv_heads", [(16, 8), (32, 8), (16, 4)])
@pytest.mark.parametrize("request_num", [1, 6])
@pytest.mark.parametrize("beam_size", [128, 256, 512])
@pytest.mark.parametrize("kv_seqlen", [128, 256, 512, 1024])
@pytest.mark.parametrize("unshared_seqlen", [2, 4])
def test_x_attention_v2_cmp(dtype, num_head, kv_heads, request_num, beam_size, kv_seqlen, unshared_seqlen):
    try:
        torch_npu.npu.set_device(int(os.environ.get("ASCEND_DEVICE_ID", 0)))
    except Exception as e:
        pytest.skip(f"NPU device not available: {e}")

    cache_key = _golden_cache_key(dtype, num_head, kv_heads, request_num, beam_size, kv_seqlen, unshared_seqlen)
    cached = _load_golden(cache_key)
    if cached is None:
        pytest.skip(f"Golden cache not found: {cache_key}. Run test_x_attention.py first.")

    v1_out = _call_op(custom_ops.x_attention_npu, cached)
    torch.npu.synchronize()

    v2_out = _call_op(custom_ops.x_attention_v2_npu, cached)
    torch.npu.synchronize()

    golden = cached["final_true_out"]

    v1_np = v1_out.cpu().float().numpy()
    v2_np = v2_out.cpu().float().numpy()
    golden_np = golden.cpu().numpy()

    cmp = precision_compare_triple_data(v2_np, v1_np, golden_np, "bf16")

    dtype_name = _DTYPE_NAME_MAP.get(dtype, str(dtype))
    _RESULTS.append({
        "dtype": dtype_name,
        "batch": request_num,
        "q_head_num": num_head,
        "kv_head_num": kv_heads,
        "beam_size": beam_size,
        "prompt_len": kv_seqlen,
        "unshared_kv_len": unshared_seqlen,
        "v1_MARE": cmp.get("v1_mare", 0.0),
        "v1_MERE": cmp.get("v1_mere", 0.0),
        "v1_RMSE": cmp.get("v1_rmse", 0.0),
        "v1_small_err": cmp.get("v1_small_err", 0),
        "v1_small_total": cmp.get("small_total", 0),
        "v2_MARE": cmp.get("v2_mare", 0.0),
        "v2_MERE": cmp.get("v2_mere", 0.0),
        "v2_RMSE": cmp.get("v2_rmse", 0.0),
        "v2_small_err": cmp.get("v2_small_err", 0),
        "v2_small_total": cmp.get("small_total", 0),
        "mare_ratio": cmp.get("mare_ratio", 0.0),
        "mere_ratio": cmp.get("mere_ratio", 0.0),
        "rmse_ratio": cmp.get("rmse_ratio", 0.0),
        "small_ratio": cmp.get("small_ratio", 0.0),
        "result": cmp.get("result", Result.ERROR).value if isinstance(cmp.get("result"), Result) else str(cmp.get("result", "")),
        "v2_mare_better": cmp.get("mare_ratio", 1.0) < 1.0,
        "v2_mere_better": cmp.get("mere_ratio", 1.0) < 1.0,
        "v2_rmse_better": cmp.get("rmse_ratio", 1.0) < 1.0,
        "v2_small_better": cmp.get("small_ratio", 1.0) < 1.0,
    })
