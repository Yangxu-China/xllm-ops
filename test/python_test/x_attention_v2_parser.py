#!/usr/bin/env python3
# coding: utf-8
# Copyright 2026 The xLLM Authors. All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     https://github.com/xLLM-AI/xllm/blob/main/LICENSE
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ==============================================================================

"""Parse and sort x_attention_v2 profiling CSV (merged or raw op_summary).

For built-in mode (raw op_summary.csv with Input Shapes):
  Extracts shape params from Input Shapes / Input Data Types columns.

For JIT mode (merged CSV from x_attention_v2_jit_merge_prof.py):
  Shape params already present as Hq, Hkv, batch, beam, prompt, maxDs, dstep.

In both cases, sorts by dtype, batch, q_head_num, kv_head_num, beam_size,
max_decode_step, prompt_len and generates an xlsx with original + sorted sheets.

Usage:
  # built-in raw op_summary
  python3 x_attention_v2_parser.py <path_to_op_summary.csv>

  # JIT merged CSV
  python3 x_attention_v2_parser.py <path_to_merged_summary.csv>

Output:
  <same_dir>/<input_name>_parsed.csv
  <same_dir>/<input_name>.xlsx
"""

import csv
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


_DTYPE_MAP = {
    "DT_BF16": "bf16",
    "DT_FLOAT16": "fp16",
    "DT_FLOAT": "fp32",
}


def _parse_input_shapes(shapes_str):
    val = shapes_str.strip().strip('"')
    entries = val.split(";")
    result = []
    for e in entries:
        e = e.strip()
        if e:
            result.append([int(x) for x in e.split(",")])
        else:
            result.append(None)
    return result


def _parse_input_dtypes(dtypes_str):
    val = dtypes_str.strip().strip('"')
    return val.split(";")


def _extract_params(input_shapes, input_dtypes):
    query = input_shapes[0]
    shared_key = input_shapes[1]
    unshared_key = input_shapes[3]
    unshared_bt = input_shapes[5]

    dtype_raw = input_dtypes[0] if input_dtypes else "DT_BF16"
    dtype = _DTYPE_MAP.get(dtype_raw, dtype_raw)

    batch = unshared_bt[0]
    q_head_num = query[1]
    kv_head_num = shared_key[1]
    beam_size = unshared_key[1]
    max_decode_step = unshared_key[3]

    if len(shared_key) == 3:
        num_shared_kv = shared_key[0]
        prompt_len = num_shared_kv // batch if batch > 0 else -1
    else:
        prompt_len = -1

    unshared_kv_len = max_decode_step

    return {
        "dtype": dtype,
        "batch": batch,
        "q_head_num": q_head_num,
        "kv_head_num": kv_head_num,
        "beam_size": beam_size,
        "max_decode_step": max_decode_step,
        "prompt_len": prompt_len,
        "unshared_kv_len": unshared_kv_len,
    }


def _is_merged_csv(fieldnames):
    """Check if CSV already has shape columns (from merge_prof)."""
    return "Hq" in fieldnames and "Hkv" in fieldnames


def _csv_to_sheet(wb, sheet_name, csv_path):
    ws = wb.create_sheet(title=sheet_name)
    with open(csv_path, newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)
    return ws


def _csv_to_sorted_sheet(wb, sheet_name, csv_path, sort_columns):
    ws = wb.create_sheet(title=sheet_name)
    with open(csv_path, newline="") as f:
        reader = csv.reader(f)
        all_rows = list(reader)
    header = all_rows[0]
    data_rows = all_rows[1:]
    col_indices = {name: i for i, name in enumerate(header)}
    sort_keys = [(col_indices[c],) for c in sort_columns if c in col_indices]

    def _sort_key(row):
        return tuple(
            (row[idx],) if not row[idx].lstrip("-").isdigit() else (int(row[idx]),)
            for idx, in sort_keys
        )

    data_rows.sort(key=_sort_key)
    ws.append(header)
    for row in data_rows:
        ws.append(row)
    return ws


def _auto_fit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_len = len(str(cell.value)) if cell.value is not None else 0
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[col_letter].width = max_len + 2


# Sort columns for both built-in (parsed) and JIT (merged) CSVs.
# Built-in uses: dtype, batch, q_head_num, kv_head_num, beam_size, max_decode_step, prompt_len
# JIT merged uses: dtype, batch, Hq, Hkv, beam, maxDs, prompt
SORT_COLUMNS_BUILTIN = [
    "dtype", "batch", "q_head_num", "kv_head_num",
    "beam_size", "max_decode_step", "prompt_len",
]
SORT_COLUMNS_JIT = [
    "dtype", "batch", "Hq", "Hkv", "beam", "maxDs", "prompt",
]


def parse_csv(input_csv_path):
    output_csv_path = input_csv_path.replace(".csv", "_parsed.csv")

    with open(input_csv_path, newline="") as f:
        reader = csv.DictReader(f)
        original_fieldnames = reader.fieldnames
        rows = list(reader)

    is_merged = _is_merged_csv(original_fieldnames)

    if is_merged:
        # JIT merged CSV: shape already present, just normalize dtype field
        for row in rows:
            dt = row.get("dtype", "")
            if dt.startswith("torch."):
                dt = dt.replace("torch.", "")
            row["dtype"] = dt
        output_fieldnames = original_fieldnames
        sort_columns = SORT_COLUMNS_JIT
    else:
        # Built-in raw op_summary: extract shape from Input Shapes
        new_columns = [
            "dtype", "batch", "q_head_num", "kv_head_num",
            "beam_size", "max_decode_step", "prompt_len", "unshared_kv_len",
        ]
        rest_columns = [c for c in original_fieldnames if c not in new_columns and c != "Task Duration(us)"]
        output_fieldnames = new_columns + ["Task Duration(us)"] + rest_columns
        for row in rows:
            input_shapes = _parse_input_shapes(row.get("Input Shapes", ""))
            input_dtypes = _parse_input_dtypes(row.get("Input Data Types", ""))
            if input_shapes and input_shapes[0] is not None:
                params = _extract_params(input_shapes, input_dtypes)
                row.update(params)
        sort_columns = SORT_COLUMNS_BUILTIN

    with open(output_csv_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=output_fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    print(f"Parsed {len(rows)} rows -> {output_csv_path}")

    xlsx_path = input_csv_path.replace(".csv", ".xlsx")
    wb = Workbook()
    wb.remove(wb.active)
    ws_orig = _csv_to_sheet(wb, "original", input_csv_path)
    ws_parsed = _csv_to_sorted_sheet(wb, "parsed", output_csv_path, sort_columns)
    _auto_fit_columns(ws_orig)
    _auto_fit_columns(ws_parsed)
    wb.save(xlsx_path)
    print(f"Xlsx saved -> {xlsx_path}")

    return output_csv_path


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(f"Usage: python3 {sys.argv[0]} <path_to_op_summary.csv or merged_summary.csv>")
        sys.exit(1)
    csv_path = sys.argv[1]
    if not os.path.exists(csv_path):
        print(f"Error: file not found: {csv_path}")
        sys.exit(1)
    parse_csv(csv_path)
